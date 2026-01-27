"""
Vistas REST responsables de entregar datos al frontend de React.
"""
import csv
import io
from datetime import datetime

from django.contrib.auth.models import User
from django.db import models, transaction
from django.http import HttpResponse
from rest_framework import generics, permissions, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.throttling import ScopedRateThrottle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.models import (
    CatalogoDispositivo,
    InstanciaDispositivo,
    Obra,
    Proyecto,
    AtributoMaestro,
    Cliente,
    Marca,
    Categoria,
    FuncionDispositivo,
    ServiciosProyecto,
    UrlsExternasProyecto,
    AtributoInstancia,
)
from core.serializers import (
    AtributoMaestroSerializer,
    CatalogoDispositivoSerializer,
    InstanciaDispositivoSerializer,
    ObraSerializer,
    ProyectoSerializer,
    RegistroUsuarioSerializer,
    ClienteSerializer,
    MarcaSerializer,
    CategoriaSerializer,
    FuncionDispositivoSerializer,
    ServiciosProyectoSerializer,
    UrlsExternasProyectoSerializer,
)
from core.services.ai_importer import extract_text_from_pdf, parse_catalog_with_gemini


def deep_clone_project(source_project_id: int, target_obra_id: int, user, nuevo_nombre: str | None = None):
    """
    Realiza una copia profunda de un Proyecto y todas sus dependencias clave (Instancias, Servicios, URLs).
    Asigna el nuevo proyecto a la obra destino y al usuario que lo clona.
    """
    try:
        source_project = Proyecto.objects.get(pk=source_project_id)
        target_obra = Obra.objects.get(pk=target_obra_id)
    except (Proyecto.DoesNotExist, Obra.DoesNotExist):
        return None

    with transaction.atomic():
        source_project.pk = None
        source_project.id = None
        source_project.obra = target_obra
        source_project.usuario_creador = user
        source_project.usuario_modificador = user
        if nuevo_nombre:
            source_project.nombre_proyecto = nuevo_nombre
        else:
            source_project.nombre_proyecto = f"{source_project.nombre_proyecto} (COPIA)"
        source_project.save()
        new_project = source_project

        for source_instance in InstanciaDispositivo.objects.filter(proyecto_id=source_project_id):
            original_funciones_usadas = list(source_instance.funciones_usadas.values_list("id", flat=True))

            # 1) Snapshot: atributos propios de la instancia (nuevo sistema)
            original_atributos = list(
                AtributoInstancia.objects.filter(instancia=source_instance).values("atributo_id", "valor")
            )

            # 2) Fallback legacy: si no hay atributos locales, copiar especificaciones del catálogo
            if not original_atributos:
                specs = source_instance.catalogo.especificaciones_set.all()
                for spec in specs:
                    if spec.valor is not None:
                        original_atributos.append(
                            {
                                "atributo_id": spec.atributo.id,
                                "valor": spec.valor,
                            }
                        )

            source_instance.pk = None
            source_instance.id = None
            source_instance.proyecto = new_project
            source_instance.usuario_creador = user
            source_instance.usuario_modificador = user
            source_instance.save()
            source_instance.funciones_usadas.set(original_funciones_usadas)

            for attr in original_atributos:
                AtributoInstancia.objects.create(
                    instancia=source_instance,
                    atributo_id=attr["atributo_id"],
                    valor=attr["valor"],
                )

        for source_servicio in ServiciosProyecto.objects.filter(proyecto_id=source_project_id):
            source_servicio.pk = None
            source_servicio.id = None
            source_servicio.proyecto = new_project
            source_servicio.obra = target_obra
            source_servicio.save()

        for source_url in UrlsExternasProyecto.objects.filter(proyecto_id=source_project_id):
            source_url.pk = None
            source_url.id = None
            source_url.proyecto = new_project
            source_url.save()

        return new_project


class AuditoriaMixin:
    """Mixin para asignar usuario creador y modificador automaticamente."""

    def perform_create(self, serializer):
        return serializer.save(
            usuario_creador=self.request.user,
            usuario_modificador=self.request.user,
        )

    def perform_update(self, serializer):
        return serializer.save(usuario_modificador=self.request.user)


class EstadoSaludAPIView(APIView):
    """Endpoint simple para revisar que la API responde correctamente."""

    permission_classes = [AllowAny]  # Permite acceso publico a esta vista

    def get(self, request, *args, **kwargs):
        """Entrega un mensaje corto para monitoreo."""
        data = {"mensaje": "API VOLTIA en linea", "total_proyectos": Proyecto.objects.count()}
        return Response(data)


class ImportarCatalogoIAView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, format=None):
        if "file" not in request.FILES:
            return Response({"error": "No se envio ningun archivo PDF"}, status=400)

        pdf_file = request.FILES["file"]

        try:
            texto_crudo = extract_text_from_pdf(pdf_file)
            items = parse_catalog_with_gemini(texto_crudo)
            return Response(
                {
                    "message": "Exito",
                    "cantidad": len(items),
                    "resultados": items,
                }
            )
        except Exception as e:
            print(f"Error IA: {e}")
            return Response({"error": str(e)}, status=500)


class ProyectoListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Expone un listado de proyectos y permite crear nuevos registros.
    Ideal para validar la escritura en PostgreSQL desde el front.
    """

    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ProyectoSerializer

    def get_queryset(self):
        """
        Filtra por obra cuando llegue el parametro ?obra=ID para reusar desde el admin.
        """
        queryset = Proyecto.objects.all().order_by("-fecha_creacion")
        obra_id = self.request.query_params.get("obra")
        if obra_id:
            queryset = queryset.filter(obra_id=obra_id)
        return queryset

class ProyectoDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateAPIView):
    """Permite consultar o actualizar un proyecto especifico."""

    queryset = Proyecto.objects.all()
    serializer_class = ProyectoSerializer
    permission_classes = [permissions.IsAuthenticated]


class RegistroUsuarioAPIView(generics.CreateAPIView):
    """
    Endpoint de API para que nuevos usuarios puedan registrarse.
    Es de solo creacion (POST).
    """

    queryset = User.objects.all()
    serializer_class = RegistroUsuarioSerializer
    permission_classes = [AllowAny]  # -íImportante! Permite el acceso sin token


class ObraListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Obras.
    Solo usuarios autenticados pueden acceder.
    """

    serializer_class = ObraSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """
        Filtra por cliente cuando llegue el parametro ?cliente=ID.
        """
        queryset = Obra.objects.all().order_by("-id")
        cliente_id = self.request.query_params.get("cliente")
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
        return queryset

class ObraDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateAPIView):
    """Permite leer o actualizar una obra puntual."""

    queryset = Obra.objects.all()
    serializer_class = ObraSerializer
    permission_classes = [permissions.IsAuthenticated]


class InstanciaDispositivoListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Instancias de Dispositivos.
    Solo usuarios autenticados pueden acceder.
    """

    serializer_class = InstanciaDispositivoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """
        Filtra por proyecto cuando llegue el parametro ?proyecto=ID.
        """
        queryset = InstanciaDispositivo.objects.all().order_by("-id")
        proyecto_id = self.request.query_params.get("proyecto")
        if proyecto_id:
            queryset = queryset.filter(proyecto_id=proyecto_id)
        return queryset

    def perform_create(self, serializer):
        """Asigna usuario creador/modificador y ajusta precio real si falta."""
        instancia = super().perform_create(serializer)
        if not instancia.precio_real or instancia.precio_real == 0:
            precio_catalogo = instancia.catalogo.precio_actual
            instancia.precio_real = precio_catalogo if precio_catalogo is not None else instancia.catalogo.precio_historico
            instancia.save(update_fields=["precio_real"])
        return instancia


class InstanciaDispositivoDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateDestroyAPIView):
    """Permite consultar, actualizar o eliminar una instancia de dispositivo puntual."""

    queryset = InstanciaDispositivo.objects.all()
    serializer_class = InstanciaDispositivoSerializer
    permission_classes = [permissions.IsAuthenticated]


class CatalogoDispositivoListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Dispositivos del Catalogo.
    """

    serializer_class = CatalogoDispositivoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """
        Permite filtrar el cat+ílogo por marca, categor+¡a o texto en nombre/modelo.
        """
        queryset = CatalogoDispositivo.objects.all().order_by("-id")
        marca_id = self.request.query_params.get("marca")
        categoria_id = self.request.query_params.get("categoria")
        search = self.request.query_params.get("q")

        if marca_id:
            queryset = queryset.filter(marca_id=marca_id)
        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)
        if search:
            queryset = queryset.filter(
                models.Q(nombre_completo_producto__icontains=search)
                | models.Q(modelo__icontains=search)
            )
        return queryset


class CatalogoDispositivoDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateDestroyAPIView):
    """
    Vista para LEER, ACTUALIZAR y BORRAR un dispositivo espec+¡fico del cat+ílogo.
    """

    queryset = CatalogoDispositivo.objects.all()
    serializer_class = CatalogoDispositivoSerializer
    permission_classes = [permissions.IsAuthenticated]


class AtributoMaestroListCreateAPIView(generics.ListCreateAPIView):
    """
    Vista para LISTAR y CREAR Atributos Maestros.
    Esta es la base para construir formularios dinamicos y filtros.
    """

    queryset = AtributoMaestro.objects.all().order_by("nombre")
    serializer_class = AtributoMaestroSerializer
    permission_classes = [permissions.IsAuthenticated]


class AtributoMaestroDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    """Vista para LEER, ACTUALIZAR y BORRAR un atributo maestro especifico."""

    queryset = AtributoMaestro.objects.all()
    serializer_class = AtributoMaestroSerializer
    permission_classes = [permissions.IsAuthenticated]


class ClienteListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Clientes.
    Solo usuarios autenticados pueden acceder.
    """

    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated]


class ClienteDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateAPIView):
    """Permite consultar o actualizar un cliente en particular."""

    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated]


class MarcaListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Marcas.
    """

    queryset = Marca.objects.all()
    serializer_class = MarcaSerializer
    permission_classes = [permissions.IsAuthenticated]


class MarcaDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateAPIView):
    """Permite consultar o actualizar una marca."""

    queryset = Marca.objects.all()
    serializer_class = MarcaSerializer
    permission_classes = [permissions.IsAuthenticated]


class CategoriaListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Categorias.
    """

    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [permissions.IsAuthenticated]


class CategoriaDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateAPIView):
    """Permite consultar o actualizar una categoria."""

    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [permissions.IsAuthenticated]


class FuncionDispositivoListCreateAPIView(AuditoriaMixin, generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Funciones de Dispositivos.
    """

    queryset = FuncionDispositivo.objects.all()
    serializer_class = FuncionDispositivoSerializer
    permission_classes = [permissions.IsAuthenticated]


class FuncionDispositivoDetailAPIView(AuditoriaMixin, generics.RetrieveUpdateAPIView):
    """Permite consultar o actualizar una funcion de dispositivo."""

    queryset = FuncionDispositivo.objects.all()
    serializer_class = FuncionDispositivoSerializer
    permission_classes = [permissions.IsAuthenticated]


class ServiciosProyectoListCreateAPIView(generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) Servicios de un Proyecto.
    """

    serializer_class = ServiciosProyectoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtra servicios por el ID del proyecto."""
        queryset = ServiciosProyecto.objects.all()
        proyecto_id = self.request.query_params.get("proyecto")
        if proyecto_id:
            queryset = queryset.filter(proyecto_id=proyecto_id)
        return queryset


class UrlsExternasProyectoListCreateAPIView(generics.ListCreateAPIView):
    """
    Vista para LISTAR (GET) y CREAR (POST) URLs externas de un Proyecto.
    """

    serializer_class = UrlsExternasProyectoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtra URLs por el ID del proyecto."""
        queryset = UrlsExternasProyecto.objects.all()
        proyecto_id = self.request.query_params.get("proyecto")
        if proyecto_id:
            queryset = queryset.filter(proyecto_id=proyecto_id)
        return queryset


class ProyectoCloneAPIView(APIView):
    """
    Endpoint para clonar un proyecto existente y todas sus dependencias.
    Requiere project_id y target_obra_id en el payload POST.
    """

    permission_classes = [permissions.IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "clonar_proyecto"

    def post(self, request, *args, **kwargs):
        source_project_id = request.data.get("source_project_id")
        target_obra_id = request.data.get("target_obra_id")
        nuevo_nombre = request.data.get("nuevo_nombre")

        if not source_project_id or not target_obra_id:
            return Response(
                {"error": "Debe proporcionar el ID del proyecto origen y el ID de la obra destino."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        new_project = deep_clone_project(source_project_id, target_obra_id, request.user, nuevo_nombre)

        if new_project:
            serializer = ProyectoSerializer(new_project)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response({"error": "No se encontr+¦ el proyecto u obra destino."}, status=status.HTTP_404_NOT_FOUND)



class ExportarMaterialesAPIView(APIView):
    """
    Genera un Excel (XLSX) multipestaña:
    - Hoja 1: Consolidado de toda la Obra.
    - Hojas N: Detalle por cada Proyecto individual.
    - Agrupación: Por atributos REALES de la instancia (Snapshot).
    """
    permission_classes = [IsAuthenticated]
    throttle_classes = [ScopedRateThrottle]
    throttle_scope = "exportar_materiales"

    def get(self, request, obra_id):
        try:
            obra = Obra.objects.get(pk=obra_id)
        except Obra.DoesNotExist:
            return Response({"error": "Obra no encontrada."}, status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        
        # --- ESTILOS ---
        font_bold = Font(bold=True)
        font_title = Font(size=14, bold=True)
        fill_header = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def generar_hoja(ws, titulo, instancias_qs, es_consolidado=False):
            ws.title = titulo[:30] # Limitación de Excel
            
            # 1. Agrupar Datos
            agregado = {}
            for inst in instancias_qs:
                cat = inst.catalogo
                
                # --- LÓGICA DE INSTANCIA (SNAPSHOT) ---
                specs_list = []
                for attr_inst in inst.atributos_set.all():
                    if attr_inst.valor:
                        nombre = attr_inst.atributo.nombre
                        unidad = attr_inst.atributo.unidad or ''
                        specs_list.append(f"{nombre}: {attr_inst.valor}{unidad}")
                specs_list.sort()
                specs_str = " - ".join(specs_list)
                
                descripcion_full = cat.nombre_completo_producto
                if specs_str:
                    descripcion_full += f" ({specs_str})"

                links_planos = ""
                if not es_consolidado:
                    planos = inst.proyecto.urls_externas.all()
                    links = [u.url for u in planos if "plano" in u.tipo_enlace.lower() or "plano" in (u.descripcion or "").lower()]
                    if not links and planos.exists():
                        links = [planos.first().url]
                    links_planos = "\n".join(links)

                key = (cat.id, descripcion_full, cat.marca.nombre if cat.marca else 'N/A')
                
                if key not in agregado:
                    agregado[key] = {
                        'cantidad': 0,
                        'modelo': cat.modelo,
                        'descripcion': descripcion_full, 
                        'marca': cat.marca.nombre if cat.marca else 'N/A',
                        'planos': links_planos,
                        'precio_historico': cat.precio_historico,
                        'precio_real_set': set(),
                    }
                agregado[key]['cantidad'] += 1
                agregado[key]['precio_real_set'].add(inst.precio_real)

            ws['A1'] = "VOLTIA LISTA DE MATERIALES"
            ws['A1'].font = font_title
            ws['F1'] = f"OBRA: {obra.nombre_obra}"
            ws['F1'].font = font_bold
            
            ws['E3'] = f"VISTA: {titulo}"
            ws['G4'] = f"FECHA: {datetime.now().strftime('%d-%m-%Y')}"
            headers = [
                "CANTIDAD",
                "FABRICANTE",
                "MODELO",
                "DESCRIPCION / ESPECIFICACIONES",
                "LINK PLANO",
                "PRECIO HISTORICO (REF)",
                "PRECIO REAL (COMPRA)",
            ]
            
            ws.append([])
            ws.append(headers)
            
            for col_num, cell in enumerate(ws[6], 1):
                cell.font = font_bold
                cell.fill = fill_header
                cell.alignment = alignment_center
                cell.border = border_thin

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 70 
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 18
            for item in agregado.values():
                precios = list(item['precio_real_set'])
                precio_real_valor = precios[0] if len(precios) == 1 else "VARIOS"
                row = [
                    item['cantidad'],
                    item['marca'],
                    item['modelo'],
                    item['descripcion'],
                    item['planos'],
                    float(item['precio_historico']) if item['precio_historico'] is not None else 0,
                    float(precio_real_valor) if precio_real_valor != "VARIOS" else precio_real_valor,
                ]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.border = border_thin
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws_total = wb.active
        
        todas_instancias = InstanciaDispositivo.objects.filter(proyecto__obra=obra)\
            .select_related('catalogo', 'catalogo__marca', 'proyecto')\
            .prefetch_related(
                'atributos_set',
                'atributos_set__atributo',
                'proyecto__urls_externas'
            )
        
        generar_hoja(ws_total, "TOTAL OBRA", todas_instancias, es_consolidado=True)

        mis_proyectos = obra.proyectos.all().order_by('id') 
        
        for proy in mis_proyectos:
            instancias_proy = todas_instancias.filter(proyecto=proy)
            if instancias_proy.exists():
                ws_proy = wb.create_sheet(title=f"P.{proy.id} - {proy.nombre_proyecto}")
                generar_hoja(ws_proy, proy.nombre_proyecto, instancias_proy, es_consolidado=False)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Materiales_{obra.nombre_obra.replace(" ", "_")}.xlsx"'
        return response

